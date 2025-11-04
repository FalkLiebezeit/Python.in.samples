import pandas as pd
from openpyxl import load_workbook

# --- Load Excel file and DataFrame ---
file_path = "./TestData/Basic_pie_chart.xlsx"
df = pd.read_excel(file_path, sheet_name="Sheet1")

# --- Open workbook and select active worksheet ---
wb = load_workbook(file_path)
ws = wb.active

# --- Set a header in cell C5 ---
ws.cell(row=5, column=3, value="Daten:")

# --- Write a list of numbers (1 to 5) into column C, starting from row 8 ---
numbers = list(range(1, 6))
for i, number in enumerate(numbers, start=1):
    ws.cell(row=i + 7, column=3, value=number)

# --- Save DataFrame back to Excel (overwrites the sheet) ---
# Note: This will overwrite the sheet, so any manual changes above will be lost.
with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    df.to_excel(writer, index=False, sheet_name="Sheet1")

# --- Save workbook changes (header and numbers) ---
wb.save(file_path)

print("\n\nExcel file was successfully updated.\n")