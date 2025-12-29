"""
Ki:

Write a Python program:

- The program loads the Excel file "Stockprice.xlsx" from the ".\TestData\" subdirectory.

- The program then saves the file in the ".\TestData\" directory.

answer: 



"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

input_file = "./TestData/Stockprices.xlsx"
output_file = "./TestData/Stockprices_saved.xlsx"

# Lade die Excel-Datei
wb = load_workbook(input_file)

ws = wb.active

# Fettgedruckte Formatierung für Zellen A1 bis D1 setzen
bold_font = Font(bold=True)
for col in ['A', 'B', 'C', 'D']:
    ws[f"{col}1"].font = bold_font


# Beispiel: Gelber Hintergrund in Zelle B2
#ws["B2"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


# Hellgraue Hintergrundfarbe für Zellen A3 bis A12 setzen
gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
for row in range(3, 13):
    ws[f"A{row}"].fill = gray_fill


# Aktienformat für Zellen B3 bis B12 setzen
stock_format = '"$"#,##0.00'  # Beispiel für US-Dollar-Formatierung
for row in range(3, 13):
    ws[f"B{row}"].number_format = stock_format


# Speichere die Datei mit Formatierungen
wb.save(output_file)

print(f"File saved successfully to {output_file}")
