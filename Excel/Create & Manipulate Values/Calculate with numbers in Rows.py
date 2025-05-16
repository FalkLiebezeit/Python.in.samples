import pandas as pd
from openpyxl import Workbook


# Eine Liste mit 10 Zahlen erstellen
zahlen = list(range(1, 11))

# Eine neue Excel-Arbeitsmappe erstellen
wb = Workbook()
ws = wb.active

# Zahlen in Spalte C (Index 3) schreiben
for i, zahl in enumerate(zahlen, start=1):
    ws.cell(row=i, column=3, value=zahl)  # Spalte C entspricht Index 3

print("Eingelesene Zahlen aus C1 bis C10:", zahlen)
#print("Zahle 1: " ws.cell("C1"))
print("Zahle 1:", ws.cell(row=1, column=3).value)


"""
# Excel-Datei laden
datei = "zahlen.xlsx"
df = pd.read_excel(datei, usecols="C")  # Nur Spalte C einlesen

# Die ersten 10 Werte aus Spalte C abrufen
zahlen = df.iloc[:10, 0].tolist()

print("Eingelesene Zahlen aus C1 bis C10:", zahlen)
"""


# Datei speichern
datei = "C:\\Users\\Falk\\TestData\\Kalkulieren.xlsx"

wb.save(datei)

print(f"Die Datei '{datei}' wurde erfolgreich erstellt mit den Zahlen in Spalte C!")

    


  
