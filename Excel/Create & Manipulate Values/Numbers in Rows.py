import pandas as pd
from openpyxl import Workbook


# Eine Liste mit 10 Zahlen erstellen
zahlen = list(range(0, 29))

# Eine neue Excel-Arbeitsmappe erstellen
wb = Workbook()
ws = wb.active

# Zahlen in Spalte C (Index 3) schreiben
for i, zahl in enumerate(zahlen, start=1):
    ws.cell(row=i, column=3, value=zahl)  # Spalte C entspricht Index 3

print("Eingelesene Zahlen aus C0 bis Xxx:", zahlen)
print("Zahle 1:", ws.cell(row=1, column=3).value)

# Datei speichern
datei = "Numbers in Rows.xlsx"


wb.save(datei)

print(f"Die Datei '{datei}' wurde erfolgreich erstellt mit den Zahlen in Spalte C!")

    


  
