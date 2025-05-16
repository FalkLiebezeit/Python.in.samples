import pandas as pd
from openpyxl import Workbook
import random


# Eine Liste mit 10 Zahlen erstellen
zahlen = list(range(1, 100))

# Eine neue Excel-Arbeitsmappe erstellen
wb = Workbook()
ws = wb.active

"""
random_integer = random.randint(1, 100)  # Returns a random integer between 1 and 100
random_float = random.uniform(1.0, 10.0)  # Returns a random float between 1.0 and 10.0
random_small = random.random()  # Returns a float between 0 and 1
"""

# Zahlen in Spalte C (Index 3) schreiben
for i, zahl in enumerate(zahlen, start=1):
    ws.cell(row=i, column=1, value=zahl)  # Spalte A entspricht Index 1
    ws.cell(row=i, column=3, value=zahl * 2) 
    ws.cell(row=i, column=5, value=random.randint(1, 100)) 
    
# Datei speichern
datei = "C:\\Users\\Falk\\TestData\\Rows with random numbers.xlsx"

wb.save(datei)

print(f"Die Datei '{datei}' wurde erfolgreich erstellt.")

    


  
