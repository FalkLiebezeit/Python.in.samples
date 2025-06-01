"""
Ki:

Write a Python program. 
The program should generate 15 integers. 
These should be stored in column 3 of an Excel spreadsheet. 
The squares of the numbers should be stored in column 5 of this Excel spreadsheet, and the spreadsheet should be saved. 
Optimize and comment on the program.

"""

import random
from openpyxl import Workbook

# Create a new workbook and select the active sheet
wb = Workbook()
ws = wb.active

# Generate 15 random integers
numbers = [random.randint(1, 100) for _ in range(15)]

# Store numbers in column 3 and their squares in column 5
for row, num in enumerate(numbers, start=1):
    ws.cell(row=row, column=3, value=num)  # Store the integer
    ws.cell(row=row, column=5, value=num ** 2)  # Store its square

# Save the workbook
wb.save(".\Testdata\generated_numbers.xlsx")

print("Excel file 'generated_numbers.xlsx' has been saved successfully.\n")