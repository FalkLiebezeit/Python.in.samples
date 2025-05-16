import pandas as pd
from openpyxl import load_workbook

# Excel-Datei laden
datei = "C:\\Users\\Falk\\TestData\\Basic pie chart diagram.xlsx"
df = pd.read_excel(datei, sheet_name="Sheet1")


# Workbook öffnen (statt neu erstellen)
wb = load_workbook(datei)
ws = wb.active


# Überschrift setzen
ws.cell(row=5, column=3, value="Daten:")


# Liste mit Zahlen erstellen
zahlen = list(range(1, 6))

# Zahlen in Spalte C schreiben
for i, zahl in enumerate(zahlen, start=1):
    ws.cell(row=i+7, column=3, value=zahl)

# Pandas DataFrame in Excel speichern
with pd.ExcelWriter(datei, engine="openpyxl") as writer:
    df.to_excel(writer, index=False)

# Änderungen speichern
wb.save(datei)

print("\n\nExcel file was successfully updated.\n")
