import os
import random
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# --- Configuration ---
# Define the file path
#file_path = "C:\\Users\\Falk\\TestData\\SalesDistribution.xlsx"
file_path = ".\\TestData\\SalesDistribution.xlsx"

# List of sample regions
REGIONS = ["Berlin", "Hamburg", "Munich", "Cologne", "Frankfurt", "Leipzig", "Dresden"]

# --- Program Start ---
print(f"Current Directory: {os.getcwd()}\n")

# Load workbook and sheets
try:
    wb = load_workbook(file_path)
    orders_sheet = wb["Bestellungen"]
    prices_sheet = wb["Preisliste"]
except FileNotFoundError:
    print(f"Error: The file at '{file_path}' was not found.")
    exit()
except KeyError as e:
    print(f"Error: The worksheet {e} was not found.")
    exit()

# --- 1. Load price data into a dictionary for fast access (Efficiency Optimization) ---
prices_data = {}
# iter_rows with values_only is faster when only reading values
for row in prices_sheet.iter_rows(min_row=2, values_only=True):
    product_id = row[0]
    if product_id is not None:
        prices_data[product_id] = {
            'name': row[1],
            'price': row[2]
        }
print("Price list successfully loaded into memory.")

# --- 2. Delete old order data ---
# Deletes all rows starting from row 2 to make space for new data
if orders_sheet.max_row > 1:
    orders_sheet.delete_rows(2, orders_sheet.max_row)
print(f"Old data in sheet '{orders_sheet.title}' deleted.")

# --- 3. Generate new orders and process them directly (Loop Optimization) ---
num_orders = random.randint(5, 19)
grouped_totals = {region: 0 for region in REGIONS}
print(f"Generating {num_orders} new orders...")

for row_index in range(2, num_orders + 2):
    # Create random order data
    order_number = row_index - 1
    product_id = random.randint(0, 19)  # Assumption: Product IDs are in the range 0-19
    quantity = random.randint(1, 50)
    region = random.choice(REGIONS)

    # Get product data from the dictionary (fast access)
    product_info = prices_data.get(product_id, {'name': "Unknown Product", 'price': 0.0})
    product_name = product_info['name']
    price = product_info['price']

    # Calculate total price
    total_price = 0.0
    if isinstance(price, (int, float)) and isinstance(quantity, (int, float)):
        total_price = price * quantity

    # Write data to the row
    orders_sheet[f'A{row_index}'] = order_number
    orders_sheet[f'B{row_index}'] = product_id
    orders_sheet[f'C{row_index}'] = quantity
    orders_sheet[f'D{row_index}'] = region
    orders_sheet[f'E{row_index}'] = product_name
    orders_sheet[f'F{row_index}'] = price
    orders_sheet[f'G{row_index}'] = total_price
    
    # Formatting for currency and highlighting
    orders_sheet[f'F{row_index}'].number_format = '"€"#,##0.00'
    orders_sheet[f'G{row_index}'].number_format = '"€"#,##0.00'
    if total_price > 100:
        orders_sheet[f'G{row_index}'].font = Font(color="FF0000", bold=True)

    # Update total sales for the region
    if region and total_price > 0:
        grouped_totals[region] += total_price

print("Orders successfully created and processed.")

# --- 4. Create summary by region ---
summary_sheet_name = "RegionSummary"
if summary_sheet_name in wb.sheetnames:
    summary_sheet = wb[summary_sheet_name]
    if summary_sheet.max_row > 1:
        summary_sheet.delete_rows(2, summary_sheet.max_row) # Delete old data
else:
    summary_sheet = wb.create_sheet(summary_sheet_name)
    summary_sheet.append(["Region", "Total Sales"]) # Header only for a new sheet

# Insert processed data
for region, total in grouped_totals.items():
    if total > 0: # Only show regions with sales
        summary_sheet.append([region, total])

# Formatting for the summary
bold_font = Font(bold=True)
for row in summary_sheet.iter_rows(min_row=1, max_row=summary_sheet.max_row):
    for cell in row:
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row[1].number_format = '"€"#,##0.00' # Currency format for the sales total

print(f"Summary created in sheet '{summary_sheet_name}'.")

# --- 5. Save the workbook ---
try:
    wb.save(file_path)
    print(f"\nExcel file successfully saved at '{file_path}'.\n")
except PermissionError:
    print(f"\nError: Permission denied to save the file. Is '{file_path}' open?\n")