"""
Ki:
Write a Python program. The program should create an Excel list with random customer data.

The list should have 20 entries:

In the first column, the customer ID, 
in the second column, a random German first name, and in the third column, 
a random German last name. Then, a postal code, 
a street address, and a house number from Germany.
All text should be centered, and the respective column widths should be adjusted.
The file should be saved as .\\TestData\\customerlist.xlsx.

"""
import pandas as pd
import random
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

customer_number = 20  # Anzahl der Kunden

# Beispiel-Daten für zufällige Kundeneinträge
first_names = ["Lukas", "Maximilian", "Leon", "Felix", "Noah", "Elias", "Paul", "Finn", "Tim", "Jonas", "Jens", "Peter", "Hans", "Joachim"]
last_names = ["Müller", "Schmidt", "Schneider", "Fischer", "Weber", "Meyer", "Wagner", "Becker", "Schulz", "Hoffmann", "Herter", "Scheunert"]
postal_codes = ["10115", "10243", "04109", "20095", "80331", "60313", "70173", "50667", "28195", "37073"]
cities = ["Berlin", "Hamburg", "München", "Köln", "Frankfurt", "Leipzig", "Dresden"]
streets = ["Hauptstraße", "Bahnhofstraße", "Gartenstraße", "Schulstraße", "Kirchstraße", "Feldstraße",
           "Dorfstraße", "Industriestraße", "Parkstraße", "Bergstraße"]

# Zufällige Kundendaten generieren
customers = []
for i in range(customer_number):
    customer = [
        i + 1,
        random.choice(first_names),
        random.choice(last_names),
        random.choice(postal_codes),
        random.choice(cities),
        random.choice(streets),
        random.randint(1, 200)  # Zufällige Hausnummer zwischen 1 und 200
    ]
    customers.append(customer)

# Dateipfad definieren
output_dir = ".\\TestData\\"
os.makedirs(output_dir, exist_ok=True)
file_path = os.path.join(output_dir, "SalesDistribution.xlsx")

# Prüfen, ob die Datei existiert
if os.path.exists(file_path):
    wb = load_workbook(file_path)
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "Customer"

# Worksheet "Customer" prüfen oder erstellen
if "Customer" in wb.sheetnames:
    ws = wb["Customer"]
else:
    ws = wb.create_sheet("Customer")

# Lösche alte Daten
ws.delete_rows(1, ws.max_row)

# Überschreibe ab Zeile 1
ws.append(["Customer ID", "First Name", "Last Name", "Postal Code", "City", "Street Address", "House Number"])
for customer in customers:
    ws.append(customer)

# Text zentrieren
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")

# Spaltenbreiten anpassen
for col in ws.columns:
    max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
    ws.column_dimensions[col[0].column_letter].width = max_length + 2

# Datei speichern
wb.save(file_path)

print(f"\nExcel-Datei gespeichert unter '{file_path}' mit zentriertem Text und angepassten Spaltenbreiten.\n")