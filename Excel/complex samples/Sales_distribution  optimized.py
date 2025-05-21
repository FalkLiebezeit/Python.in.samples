from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, numbers
import random
import os

# Define the file path
file_path = "C:\\Users\\Falk\\TestData\\SalesDistribution.xlsx"

# Load the workbook and sheets
wb = load_workbook(file_path)
orders_sheet = wb["Bestellungen"]
prices_sheet = wb["Preisliste"]

# Generate a random number of orders
num_orders = random.randint(5, 19)

# Print the current working directory
print(f"\nCurrent Directory: {os.getcwd()}\n")

# Function to retrieve product name by ID
def get_product(product_id):
    """Returns the product name corresponding to the given ID."""
    for row in prices_sheet.iter_rows(min_row=2, max_row=prices_sheet.max_row, min_col=1, max_col=5):
        if row[0].value == product_id:
            return row[1].value
    return "Unknown Product"

# Function to retrieve price by ID
def get_price(product_id):
    """Returns the price corresponding to the given ID."""
    for row in prices_sheet.iter_rows(min_row=2, max_row=prices_sheet.max_row, min_col=1, max_col=5):
        if row[0].value == product_id:
            return row[2].value
    return 0.0  # Default price if not found

# List of example regions
regions = ["Berlin", "Hamburg", "Munich", "Cologne", "Frankfurt", "Leipzig", "Dresden"]

# Remove old orders (clear rows from 2 to 20)
orders_sheet.delete_rows(2, 19)

# Fill in order details
order_number = 1
for row in orders_sheet.iter_rows(min_row=2, max_row=num_orders, min_col=1, max_col=5):
    row[0].value = order_number  # Assign order number
    row[1].value = random.randint(0, 19)  # Random product ID
    row[2].value = random.randint(1, 50)  # Random quantity (minimum 1)
    row[3].value = random.choice(regions)  # Assign a random region
    order_number += 1

# Populate product names and prices in the "Bestellungen" sheet
for row in orders_sheet.iter_rows(min_row=2, max_row=orders_sheet.max_row, min_col=1, max_col=6):
    product_id = row[1].value
    row[4].value = get_product(product_id)  # Product Name
    row[5].value = get_price(product_id)  # Product Price

# Calculate the total price for each order
for row in orders_sheet.iter_rows(min_row=2, max_row=orders_sheet.max_row, min_col=1, max_col=7):
    price, quantity = row[5].value, row[2].value

    if price is not None and quantity is not None:
        total_price = price * quantity
        row[6].value = total_price  # Store total price
        # Highlight orders with a high total price
        if total_price > 100:
            row[6].font = Font(color="FF0000")  # Red font for high-value orders
    else:
        row[6].value = "No Data"

# **Grouping orders by region**
grouped_totals = {}

for row in orders_sheet.iter_rows(min_row=2, max_row=orders_sheet.max_row, min_col=3, max_col=7):
    region = row[0].value  # Get the region name
    total_price = row[4].value  # Get total order price
    
    if region not in grouped_totals:
        grouped_totals[region] = 0
    
    if total_price is not None and isinstance(total_price, (int, float)):
        grouped_totals[region] += total_price

# **Create a new sheet for grouping summary**
if "GroupedTotals" in wb.sheetnames:
    summary_sheet = wb["GroupedTotals"]
    summary_sheet.delete_rows(2, summary_sheet.max_row)  # Clear previous data
else:
    summary_sheet = wb.create_sheet("GroupedTotals")

# Add headers
summary_sheet.append(["Region", "Total Sales"])

# Insert grouped totals into the new sheet
for region, total in grouped_totals.items():
    summary_sheet.append([region, total])

# Apply formatting to grouped total rows
bold_font = Font(bold=True)
for row in summary_sheet.iter_rows(min_row=1, max_row=summary_sheet.max_row, min_col=1, max_col=2):
    for cell in row:
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="center")

# Save the modified workbook
wb.save(file_path)
print("Excel file saved successfully!\n")