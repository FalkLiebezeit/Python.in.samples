from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, numbers
import random

datei_pfad = "C:\\Users\\Falk\\TestData\\SalesDistribution.xlsx"
#datei_pfad ="SalesDistribution.xlsx"

wb = load_workbook(datei_pfad)

bestellungen = wb["Bestellungen"]
preisliste   = wb["Preisliste"]

anzahl_bestellungen = random.randint(5, 19) 



def get_product(id):
    for row in preisliste.iter_rows(min_row=2, max_row=preisliste.max_row , min_col=1, max_col=5):
        if row[0].value == id:
            return row[1].value
    return -1

def get_price(id):
    for row in preisliste.iter_rows(min_row=2, max_row=preisliste.max_row , min_col=1, max_col=5):
        if row[0].value == id:
            return row[2].value
    return -1


# Beispiel-Liste mit Regionen
region      = ["Berlin", "Hamburg", "München", "Köln", "Frankfurt","Hamburg", "Leipzig", "Dresden"]
bundesland  = ["Brandenburg","Sachsen","Mecklenburg","Thüringen","Sachsen-Anhalt"]



bundeslaender_mit_staedten = {
    "Baden-Württemberg": ["Stuttgart", "Karlsruhe", "Mannheim", "Freiburg", "Heidelberg", "Ulm", "Pforzheim", "Reutlingen", "Heilbronn", "Esslingen"],
    "Bayern": ["München", "Nürnberg", "Augsburg", "Regensburg", "Ingolstadt", "Fürth", "Würzburg", "Erlangen", "Bayreuth", "Bamberg"],
    "Berlin": ["Berlin"],
    "Brandenburg": ["Potsdam", "Cottbus", "Brandenburg an der Havel", "Frankfurt (Oder)", "Oranienburg", "Eberswalde", "Falkensee", "Neuruppin", "Senftenberg", "Luckenwalde"],
    "Bremen": ["Bremen", "Bremerhaven"],
    "Hamburg": ["Hamburg"],
    "Hessen": ["Frankfurt am Main", "Wiesbaden", "Kassel", "Darmstadt", "Offenbach am Main", "Hanau", "Gießen", "Marburg", "Fulda", "Limburg an der Lahn"],
    "Mecklenburg-Vorpommern": ["Schwerin", "Rostock", "Neubrandenburg", "Stralsund", "Greifswald", "Wismar", "Güstrow", "Bergen auf Rügen", "Anklam", "Ludwigslust"],
    "Niedersachsen": ["Hannover", "Braunschweig", "Oldenburg", "Osnabrück", "Göttingen", "Wolfsburg", "Hildesheim", "Salzgitter", "Celle", "Lüneburg"],
    "Nordrhein-Westfalen": ["Köln", "Düsseldorf", "Dortmund", "Essen", "Duisburg", "Bochum", "Bonn", "Münster", "Bielefeld", "Gelsenkirchen"],
    "Rheinland-Pfalz": ["Mainz", "Ludwigshafen am Rhein", "Koblenz", "Trier", "Kaiserslautern", "Worms", "Neuwied", "Frankenthal", "Zweibrücken", "Pirmasens"],
    "Saarland": ["Saarbrücken", "Neunkirchen", "Völklingen", "Homburg", "Merzig", "St. Ingbert", "Saarlouis", "Dillingen/Saar", "Blieskastel", "Lebach"],
    "Sachsen": ["Leipzig", "Dresden", "Chemnitz", "Zwickau", "Plauen", "Görlitz", "Freital", "Bautzen", "Hoyerswerda", "Riesa"],
    "Sachsen-Anhalt": ["Magdeburg", "Halle (Saale)", "Dessau-Roßlau", "Lutherstadt Wittenberg", "Stendal", "Halberstadt", "Weißenfels", "Bitterfeld-Wolfen", "Sangerhausen", "Bernburg"],
    "Schleswig-Holstein": ["Kiel", "Lübeck", "Flensburg", "Neumünster", "Norderstedt", "Elmshorn", "Pinneberg", "Itzehoe", "Schleswig", "Ahrensburg"],
    "Thüringen": ["Erfurt", "Jena", "Gera", "Weimar", "Eisenach", "Nordhausen", "Suhl", "Mühlhausen", "Gotha", "Altenburg"]
}

# Beispiel für Zugriff auf die Städte eines Bundeslands
print(bundeslaender_mit_staedten["Sachsen"])

bundeslaender_array = [
    ["Baden-Württemberg", ["Stuttgart", "Karlsruhe", "Mannheim", "Freiburg", "Heidelberg", "Ulm", "Pforzheim", "Reutlingen", "Heilbronn", "Esslingen"]],
    ["Bayern", ["München", "Nürnberg", "Augsburg", "Regensburg", "Ingolstadt", "Fürth", "Würzburg", "Erlangen", "Bayreuth", "Bamberg"]],
    # Weitere Bundesländer ...
]

print(bundeslaender_array[0][1])  # Gibt die größten Städte in Baden-Württemberg aus




# alte Bestellungen löchen
bestellungen.delete_rows(2, 19)  # Entfernt Zeilen 2 bis 20

    


nr = 1
for row in bestellungen.iter_rows(min_row=2, max_row=anzahl_bestellungen, min_col=1, max_col=5):
    row[0].value = nr
    nr += 1  # Korrekte Inkrementierung der Nummerierung
    row[1].value = random.randint(0, 19)
    row[2].value = random.randint(0, 50)
    row[3].value = region[random.randint(0, len(region) - 1)]  # Index anpassen


# Namen und Preise in Tabelle Bestellungen ausfüllen
for row in bestellungen.iter_rows(min_row=2, max_row=bestellungen.max_row, min_col=1, max_col=6):
    product_id = row[1].value
    row[4].value = get_product(product_id)
    row[5].value = get_price(product_id)


# Gesamtpreis für Bestellung berechnen
for row in bestellungen.iter_rows(min_row=2, max_row=bestellungen.max_row, min_col=1, max_col=7):

    preis, menge = row[5].value, row[2].value

    if menge is not None and preis is not None:
        row[6].value = menge * preis
        if row[6].value > 100:
            row[6].font = Font(color="FF0000")  # Rot - bei hohem Bestellwert Hintergrund der Zelle ändern
    else:
        row[6].value = "keine Angabe"




# Formatierungen

# Fettformatierung definieren
bold_font = Font(bold=True)

# Hintergrundfarbe definieren (z. B. Gelb)
fill_color = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Formatierung auf Bereich A1:G7 anwenden
for row in bestellungen.iter_rows(min_row=1, max_row=1, min_col=1, max_col=7):
    for cell in row:
        cell.font = bold_font
        cell.fill = fill_color


# Formatierung auf Bereich A2:G20 anwenden
for row in bestellungen.iter_rows(min_row=2, max_row=bestellungen.max_row, min_col=1, max_col=7):
    for cell in row:
         cell.alignment = Alignment(horizontal="center", vertical="center")

 # Setze das Zahlenformat auf Währung für den Bereich F2 bis G20
for row in bestellungen.iter_rows(min_row=2, max_row=20, min_col=6, max_col=7):  # F = 6, G = 7
    for cell in row:
        cell.number_format = numbers.FORMAT_NUMBER_00
        


wb.save(datei_pfad)
print("Excel-Datei erfolgreich gespeichert.")